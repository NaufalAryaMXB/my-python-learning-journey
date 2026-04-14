from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import os

if os.path.exists("todolist.xlsx"):
    wb = load_workbook("todolist.xlsx")
    sheet = wb.active
else:
    wb = Workbook()
    sheet = wb.active
    sheet.append(["Time", "Task", "Status"])
    wb.save("todolist.xlsx")

while True:
    print("\n========================================")
    print("Welcome  to the to do list pick an option:")
    print("1 to  add  a task to the list")
    print("2 to change  the status of  a task")
    print("3 to delete a task")
    print("4  to  show all of the task")
    print("5 to exit the program")
    print("========================================")
    option = int(input("Enter a Number: "))

    if option == 1:
        print("----------------------------------------")
        addtask = input("Add a task: ")
        now = datetime.datetime.now()
        auto_time = now.strftime("%Y-%m-%d %H:%M")
        status = "Pending"
        try:
            new_row = [auto_time, addtask, status]
            sheet.append(new_row)
            wb.save("todolist.xlsx")
            print("----------------------------------------")
        except: 
            print("An error occured")
        else : 
            print(addtask, "task added")

    elif option == 2:
        if sheet.max_row <= 1:
            print("You don't have any task")
            print("----------------------------------------")
        else:
            print("----------------------------------------")
            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                time_created = row[0]
                task_name = row[1]
                status = row[2]
                print(f"{index}. [{status}] {task_name}")
            changest = int(input("Select the Number of the task you want it's status to change to Done: "))
            target = changest
            statuscol = 3
            new_status =  "Done"
            try:
                print("----------------------------------------")
                sheet.cell(row=target+1, column=statuscol).value = new_status
                wb.save("todolist.xlsx")
            except:
                print("An  error occured")
            else:
                print("Task Status Sucessfully Chnaged")

    elif option == 3:
        if sheet.max_row <= 1:
                print("You don't have any task")
                print("----------------------------------------")
        else:
            print("----------------------------------------")
            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                time_created = row[0]
                task_name = row[1]
                status = row[2]
                print(f"{index}. [{status}] {task_name}")
            delete = int(input("Select the Number of the task you want to Delete: "))
            try :
                sheet.delete_rows(delete+1)
                wb.save("todolist.xlsx")
            except:
                print("An error occured deletion failed")
            else:
                print("Task Sucessfullt deleted")

    elif option == 4:
        if sheet.max_row <= 1:
            print("You don't have any task")
            print("----------------------------------------")
        else:
            print("----------------------------------------")
            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                time_created = row[0]
                task_name = row[1]
                status = row[2]
                print(f"{index}. [{status}] {task_name}")

    elif option == 5:
        print("----------------------------------------")
        print("Thank you for using the program")
        print("========================================\n")
        break

    else:
        print("----------------------------------------")
        print("input Unknown,Try Again")